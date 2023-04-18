import os
import shutil
from PIL import Image, ImageDraw, ImageFont, ImageColor
from textwrap3 import wrap
from openpyxl import load_workbook

try:
    from configparser import ConfigParser
except ImportError:
    from ConfigParser import ConfigParser  # ver. < 3.0

if not os.path.exists("/config/config.ini"):
    shutil.copy("/config-copy/config.ini", "/config/config.ini")
if not os.path.exists("/config/txt_generator.xlsx"):
    shutil.copy("/config-copy/txt_generator.xlsx", "/config/txt_generator.xlsx")
for font_extension in os.listdir('/config'):
    if not font_extension.endswith('.ttf'):
        shutil.copy("/config-copy/XiaolaiP-SC-Light.ttf", "/config/XiaolaiP-SC-Light.ttf")

# instantiate
config = ConfigParser()

# parse existing file
config.read('/config/config.ini')

# Excel 行數範圍設定

start_row = config.getint('Excel_Setting', 'Start_Row')
end_row = config.getint('Excel_Setting', 'End_Row')

# 圖片大小設定
WIDTH = config.getint('Image_Setting', 'Image_WIDTH')
HEIGHT = config.getint('Image_Setting', 'Image_HEIGHT')

# 字體設定
font_family = config.get('Font_Setting', 'FONT_FAMILY')
font_size = config.getint('Font_Setting', 'FONT_SIZE')
v_margin = config.getint('Font_Setting', 'V_MARGIN')
char_limit = config.getint('Font_Setting', 'CHAR_LIMIT')
text_color = config.get('Font_Setting', 'TEXT_COLOR')


def get_y_and_heights(text_wrapped, dimensions, margin, font):
    """Get the first vertical coordinate at which to draw text and the height of each line of text"""
    # https://stackoverflow.com/a/46220683/9263761
    ascent, descent = font.getmetrics()

    # Calculate the height needed to draw each line of text (including its bottom margin)
    line_heights = [
        font.getmask(text_line).getbbox()[3] + descent + margin
        for text_line in text_wrapped
    ]
    # The last line doesn't have a bottom margin
    line_heights[-1] -= margin

    # Total height needed
    height_text = sum(line_heights)

    # Calculate the Y coordinate at which to draw the first line of text
    y = (dimensions[1] - height_text) // 2

    # Return the first Y coordinate and a list with the height of each line
    return (y, line_heights)

workbook = load_workbook(filename='/config/txt_generator.xlsx')
sheet = workbook.active

# 設定要讀取的欄位位置
text_column = 'A'
color_column = 'B'

for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
    text, color = row[0], row[1]
    
    # Create the font
    font_family_path = f"/config/{font_family}"
    font = ImageFont.truetype(font_family_path, font_size)
    # New image based on the settings defined above
    img = Image.new("RGB", (WIDTH, HEIGHT), color=(0, 0, 0))

    # 設定圖片底色
    if color is not None:
        background_color = ImageColor.getrgb('#' + str(color))
    else:
        background_color = (0, 0, 0)

    # 建立一個新的底色圖片
    draw_interface = ImageDraw.Draw(img)
    draw_interface.rectangle([(0,0), (WIDTH,WIDTH)], fill=background_color)

    # Wrap the `text` string into a list of `char_limit`-character strings
    text_lines = wrap(text, char_limit)
    # Get the first vertical coordinate at which to draw text and the height of each line of text
    y, line_heights = get_y_and_heights(
        text_lines,
        (WIDTH, HEIGHT),
        v_margin,
        font
    )

    # Draw each line of text
    for i, line in enumerate(text_lines):
        # Calculate the horizontally-centered position at which to draw this line
        line_width = font.getmask(line).getbbox()[2]
        x = ((WIDTH - line_width) // 2)

        # Draw this line
        draw_interface.text((x, y), line, font=font, fill=text_color)

        # Move on to the height at which the next line should be drawn at
        y += line_heights[i]

    # 儲存圖片
    filename = f"/images/{text}.jpg"
    img.save(filename)

